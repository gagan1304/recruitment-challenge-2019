"""
Comand for comparing two sources of solar radiance data

Run from the command line like so
```
    > python process_data.py INPUT_CSV_FILE POSTCODE --output_file=OUTPUT_CSV_FILE --output_endpoint=OUTPUT_API_URL
```

Inputs:
 * CSV data from a weather station (first command line argument)

Outputs:
 * CSV to write results to (second command line argument)
 * endpoint for posting JSON results (third command line argument)

"""
import sys
#from lib import model
from lib import process_bom_grid_file as pr
import pandas as pd
import numpy as np
import os
import datetime
from os import listdir
from os.path import isfile, join
from matplotlib import pyplot
import json
from lib.model import detect_anomalies
from lib import model 
import pytz
import pysolar
import openpyxl
import requests



#%%
if __name__ == '__main__':
    
    lat = -27.5
    lon = 153.0
    # Read weather station solar data CSV (first command line argument)
    print('Reading weather station solar data...')
    # Relative filepath for weather data
    weatherDataFile = "Files/weather_station_20190601-20190731.csv"
    print(os.path.isfile(weatherDataFile)) # check if file exists   
    # Lambda function for reading excel timestamp
    dateparse = lambda x: pd.datetime.strptime(x,'%d/%m/%Y %H:%M:%S')
    # Read the weather data file
    df_ws = pd.read_csv(weatherDataFile,parse_dates=['Time'],date_parser=dateparse)
    df_ws['Solar Radiation'].fillna(0,inplace=True) # replace all nans with 0
    # Creating a Megajoule/m2 column
    df_ws['Solar Radiation MJ'] = df_ws['Solar Radiation']*(5*60)/1000000 # 5 minute interval
    # assigning index to time column 
    date_index = pd.DatetimeIndex(df_ws['Time'])
    # Sorting the data to daily values
    df_ws_daily = df_ws.groupby([date_index.year, date_index.month, date_index.day])\
                .agg({'Time':'min','Solar Radiation': 'mean','Solar Radiation MJ': 'sum'})\
                .rename(columns={'level_1': 'year',
                                 'level_2': 'month',
                                 'level_3': 'day'})
    # Sorting the data to hourly values
    df_ws_hourly = df_ws.groupby([date_index.year, date_index.month, date_index.day, date_index.hour])\
                .agg({'Time': 'min','Solar Radiation': 'mean','Solar Radiation MJ': 'sum'})\
                .rename(columns={'level_1': 'year',
                                 'level_2': 'month',
                                 'level_3': 'day',
                                 'level_4': 'hour'}) 
    

    #%% Read daily total data (from grid files / or directly from BOM?)
    print('Reading BOM solar data...')
    # Path for BOM data
    BOMDataFilePath = 'Files/BOM data/'
    # list of all the BOM files
    BOMfiles = [f for f in listdir(BOMDataFilePath) if isfile(join(BOMDataFilePath, f))]
    numDays = 61 # number of days June and July month
    BOM_daily = np.zeros((numDays,)) # list of daily BOM data (in MJ/m2)
    counter = 0
    # get daily BOM data from grid files
    for i in BOMfiles:
        grid_file_name = join(BOMDataFilePath,i)    
        data_by_location, start_date, end_date = pr.get_solar_exposure_data_from_grid_file(grid_file_name, tzinfo=None)
        print(data_by_location[str(lat)][str(lon)], start_date, end_date)
        BOM_daily[counter] =  data_by_location[str(lat)][str(lon)]
        counter = counter + 1    

    #%% Align to a hourly timestamp
    #
    
    startDateTime = datetime.datetime(2018,5,31,14,tzinfo=pytz.UTC)
    startDateTime = startDateTime.astimezone(pytz.timezone('Australia/Brisbane'))
    #hourly variables
    totalHours = 24*numDays
    alt = np.zeros((totalHours,)) # altitude array
    rad = np.zeros((totalHours,)) # radiation array
    dt_hourly_array = np.empty(totalHours, dtype='datetime64[s]')
    #dt_hourly_array2 = np.zeros((totalHours,4))
    # Using pysolar to get the hourly solar profile for June and July month
    for i in range(totalHours):
        currentDateTime = startDateTime + datetime.timedelta(hours=i)
        currentAltitude = pysolar.solar.get_altitude(lat,lon,currentDateTime)
        if currentAltitude <= 0:
            currentRadiation = 0
        else:
            currentRadiation = pysolar.radiation.get_radiation_direct(currentDateTime,currentAltitude)
        alt[i] = currentAltitude
        rad[i] = currentRadiation # hourly radiation
#        dt_hourly_array[i] = currentDateTime   
#        dt_hourly_array2[i,0] = dt_hourly_array[i].astype(datetime.datetime).year
#        dt_hourly_array2[i,1] = dt_hourly_array[i].astype(datetime.datetime).month
#        dt_hourly_array2[i,2] = dt_hourly_array[i].astype(datetime.datetime).day
#        dt_hourly_array2[i,3] = dt_hourly_array[i].astype(datetime.datetime).hour
    
    normalised_rad = np.zeros((totalHours,))
    # Normalisation to BOM data 
    for i in range(numDays):
        currentBOMValue = BOM_daily[i]
        normalisation_factor = sum(rad[i*24:(i+1)*24]*3600/1000000)/currentBOMValue
        normalised_rad[i*24:(i+1)*24] = rad[i*24:(i+1)*24]/normalisation_factor
    
    #%%

    # Write output to CSV
    # utc timestamp, solar_ws, solar_BOM
    outputFile = 'output.csv'    
    wb = openpyxl.Workbook() 
    ws = wb.active       

    ws['A1'] = 'UTC Tmestamp'
    ws['B1'] = 'solar_ws'
    ws['C1'] = 'solar_BOM'
    
    for i in range(totalHours):
        ws.cell(row=i+2,column=1).value = df_ws_hourly['Time'][i].strftime('%d/%m/%Y %H:%M')
        ws.cell(row=i+2,column=2).value = df_ws_hourly['Solar Radiation'][i]
        ws.cell(row=i+2,column=3).value = normalised_rad[i]
    
    wb.save(outputFile) 

    #%% Visualise results
    pyplot.figure(1)
    pyplot.plot(df_ws_daily['Time'],BOM_daily)
    pyplot.plot(df_ws_daily['Time'],df_ws_daily['Solar Radiation MJ'])    
    pyplot.title('Daily Solar Radiation Comparison')
    pyplot.ylabel('MJ/m2')
    pyplot.xlabel('Days')
    pyplot.legend(['BOM Data','Weather Station Data'], loc='best')    
    
    
    pyplot.figure(2)
    toRange = 24*10
    pyplot.plot(df_ws_hourly['Time'][0:toRange],normalised_rad[0:toRange])
    pyplot.plot(df_ws_hourly['Time'][0:toRange],df_ws_hourly['Solar Radiation'][0:toRange])
    #pyplot.plot(df_ws_hourly['Time'][0:toRange],range(toRange),normalised_rad[0:toRange]-df_ws_hourly['Solar Radiation'][0:toRange])
    pyplot.title('Hourly Solar Radiation Comparison')
    pyplot.ylabel('Watts/m2')
    pyplot.xlabel('Hours')
    pyplot.legend(['BOM Data','Weather Station Data'], loc='best')
    
    
    #%% Additional Task: Write output to an endpoint
    
    uploadDict = {
            "candidate" : "gagan1304",
            "version" : "final",
            }
    records = []
    for i in range(totalHours):
        records.append({"utc_timestamp":df_ws_hourly['Time'][i].isoformat(), "solar_ws":df_ws_hourly['Solar Radiation'][i], "solar_bom":normalised_rad[i]})    # to be changed to hourly
    uploadDict['records'] = records
    jsonData=json.dumps(uploadDict)
    outputfilename = 'testOut.txt'
    
    with open(outputfilename, 'w') as outfile:
        json.dump(uploadDict, outfile)
    

    url = 'https://qs3w5fq4oi.execute-api.ap-southeast-2.amazonaws.com/dev/ping'    
    headers = {'Content-type': 'application/json', 'Accept': 'text/plain'}
    r = requests.post(url, data=json.dumps(uploadDict), headers=headers)

    
    #%%
    # Additional Task: detect anomalies
    anom = detect_anomalies(df_ws_hourly['Time'].to_numpy(), normalised_rad, df_ws_hourly['Solar Radiation'].to_numpy())
#    pyplot.figure(3)
#    pyplot.plot(anom['feature'],'r',linewidth=3)
#    pyplot.plot(anom['mean_shift'],'g')    
    anomaly2 = np.zeros(anom['mean_shift'].shape)
    anomaly2.fill(np.nan)
    indexAn = np.where(anom['mean_shift']>1)
    indexAn2 = np.ones(indexAn[0].shape,)
    for i in range(indexAn[0].shape[0]):
        indexAn2[i] = int(indexAn[0][i])
        #print(indexAn2[i])
    
    anomaly2[indexAn2.astype(int)] = normalised_rad[indexAn2.astype(int)]
    #for i in range(560):
        #print(str(i) + ". Anomaly = " + str(anomaly2[i]) + " BOM =" + str(normalised_rad[i]) )
        
    pyplot.figure(4)
    toRange = 24*10
    pyplot.plot(df_ws_hourly['Time'][0:toRange],normalised_rad[0:toRange])
    pyplot.plot(df_ws_hourly['Time'][0:toRange],df_ws_hourly['Solar Radiation'][0:toRange])
    pyplot.plot(df_ws_hourly['Time'][0:toRange],anomaly2[0:toRange],'r*',linewidth=5)
    pyplot.title('Hourly Solar Radiation Comparison')
    pyplot.ylabel('Watts/m2')
    pyplot.xlabel('Hours')
    pyplot.legend(['BOM Data','Weather Station Data','Anomalies'], loc='best')
    pyplot.show()

    
    
   















    
    
    
    
    
    
    
    
    

    
    
    
    
    
    
    
    
    
    
    
    


